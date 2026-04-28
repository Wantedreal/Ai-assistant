"""
importer.py
-----------
Reads a battery-cells Excel workbook (openpyxl, no pandas) and
re-populates the local SQLite catalogue.

Cleaning rules applied on import:
- Duplicates (same Product_Number): keep row with most filled fields;
  tie-break by highest CycleLife.
- Chemistry: normalize NMC variants → 'NMC'; derive from broad Chemistry
  when Chemistry_Detail is '-'.
- 'not defined' / '-' strings → NULL.
- energie_volumique_wh_l > 800 Wh/L → NULL (physically implausible).
- dod_reference_pct stored as percentage (0–100); converts 0–1 fractions.

Config (source path for Sync) is stored alongside the DB as import_config.json.
"""

import json
import logging
from io import BytesIO
from pathlib import Path

import openpyxl
from sqlalchemy.orm import Session

from app.core.config import DATABASE_PATH
from app.models.cellule import Cellule

CONFIG_PATH = DATABASE_PATH.parent / "import_config.json"

REQUIRED_COLUMNS = {
    "nom", "longueur_mm", "largeur_mm", "hauteur_mm",
    "masse_g", "tension_nominale", "capacite_ah",
    "courant_max_a", "type_cellule", "taux_swelling_pct",
}

# New-format Excel column → DB field mapping (all optional/nullable)
EXTENDED_COLUMN_MAP = {
    "Product_Number":           "nom",
    "Cell_Format":              "type_cellule",
    "Capacity_Ah":              "capacite_ah",
    "NomVoltage_Volt":          "tension_nominale",
    "Weight_gr":                "masse_g",
    "Length_mm":                "longueur_mm",
    "Width_mm":                 "largeur_mm",
    "Thickness_Height_mm":      "hauteur_mm",
    "Diameter_mm":              "diameter_mm",
    "CycleLife":                "cycle_life",
    "t_cycle_DOD":              "dod_reference_pct",
    "Discharge_MaxConstant_Crate": "c_rate_max_discharge",
    "Charge_MaxConstant_Crate": "c_rate_max_charge",
    "LastCycle_ResidualCapacity_%": "eol_capacity_pct",
    "CutOff_Voltage_Volt":      "cutoff_voltage_v",
}

_NULL_STRINGS = {None, "-", "not defined", "not applicable", ""}


def _is_null(v) -> bool:
    if v is None:
        return True
    return str(v).strip().lower() in {"", "-", "not defined", "not applicable"}


def _to_float(v) -> float | None:
    if _is_null(v):
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def _to_int(v) -> int | None:
    if _is_null(v):
        return None
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return None


_SWELLING_BY_CHEMISTRY = {
    "LFP": 0.03,
    "NMC": 0.08,
    "NCA": 0.08,
    "LTO": 0.01,
    "LCO": 0.05,
}
_SWELLING_DEFAULT = 0.05


def _clean_chemistry(detail, broad) -> str | None:
    """Normalize chemistry to one of: NMC, LFP, NCA, LTO, LCO, or None."""
    if not _is_null(detail):
        d = str(detail).strip()
        if d.upper().startswith("NMC"):
            return "NMC"
        if d.upper() in ("LFP", "NCA", "LTO", "LCO"):
            return d.upper()
        if d.upper().startswith("LCO"):
            return "LCO"
        return d
    # Derive from broad chemistry
    if not _is_null(broad):
        b = str(broad).strip().upper()
        if b == "NI-RICH":
            return "NMC"
        if b in ("LFP", "NCA", "LTO", "LCO"):
            return b
    return None


def _completeness(row: tuple) -> int:
    """Count non-null values in a row — used to pick best duplicate."""
    return sum(1 for v in row if not _is_null(v))


def get_source_path() -> str | None:
    if CONFIG_PATH.exists():
        return json.loads(CONFIG_PATH.read_text()).get("source_path")
    return None


def save_source_path(path: str) -> None:
    CONFIG_PATH.write_text(json.dumps({"source_path": path}))


def _is_extended_format(headers: list) -> bool:
    """True if the workbook uses the new TUM/ISI extended format."""
    return "Product_Number" in headers


def _import_extended(ws, db: Session) -> int:
    """Import from new-format Excel (TUM/ISI dataset with Phase 1 fields)."""
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers) if h}

    rows = [
        row for row in ws.iter_rows(min_row=2, values_only=True)
        if not all(_is_null(v) for v in row)
    ]

    # ── Deduplicate: keep best row per Product_Number ─────────────────────────
    best: dict[str, tuple] = {}
    for row in rows:
        name = row[idx.get("Product_Number", -1)] if "Product_Number" in idx else None
        if _is_null(name):
            continue
        name = str(name).strip()
        if name not in best:
            best[name] = row
        else:
            existing = best[name]
            new_score = _completeness(row)
            old_score = _completeness(existing)
            if new_score > old_score:
                best[name] = row
            elif new_score == old_score:
                # tie-break: prefer higher CycleLife
                new_cl = _to_int(row[idx["CycleLife"]]) if "CycleLife" in idx else 0
                old_cl = _to_int(existing[idx["CycleLife"]]) if "CycleLife" in idx else 0
                if (new_cl or 0) > (old_cl or 0):
                    best[name] = row

    db.query(Cellule).delete()

    count = 0
    try:
        for name, row in best.items():
            cell_type = str(row[idx["Cell_Format"]]).strip() if "Cell_Format" in idx else "Pouch"

            # Dimensions — cylindrical uses Diameter_mm for longueur_mm
            dia = _to_float(row[idx["Diameter_mm"]]) if "Diameter_mm" in idx else None
            length = _to_float(row[idx["Length_mm"]]) if "Length_mm" in idx else None
            width = _to_float(row[idx["Width_mm"]]) if "Width_mm" in idx else None
            thick = _to_float(row[idx["Thickness_Height_mm"]]) if "Thickness_Height_mm" in idx else None

            if cell_type.lower() == "cylindrical":
                longueur = dia or length or 0.0
                largeur = dia or width or 0.0
            else:
                longueur = length or 0.0
                largeur = width or 0.0

            # dod_reference_pct: stored as 0–100; convert if value is 0–1 fraction
            dod_raw = _to_float(row[idx["t_cycle_DOD"]]) if "t_cycle_DOD" in idx else None
            if dod_raw is not None and dod_raw <= 1.0:
                dod_raw = round(dod_raw * 100.0, 1)

            db.add(Cellule(
                nom               = name,
                type_cellule      = cell_type,
                longueur_mm       = longueur,
                largeur_mm        = largeur,
                hauteur_mm        = thick or 0.0,
                diameter_mm       = dia,
                masse_g           = (_to_float(row[idx["Weight_gr"]]) if "Weight_gr" in idx else None) or 0.0,
                tension_nominale  = (_to_float(row[idx["NomVoltage_Volt"]]) if "NomVoltage_Volt" in idx else None) or 0.0,
                capacite_ah       = (_to_float(row[idx["Capacity_Ah"]]) if "Capacity_Ah" in idx else None) or 0.0,
                courant_max_a     = (
                    _to_float(row[idx["Discharge_MaxConstant_A"]]) if "Discharge_MaxConstant_A" in idx else None
                ) or (
                    # fallback: C-rate × capacity when MaxConstant_A is missing
                    (_to_float(row[idx["Discharge_MaxConstant_Crate"]]) or 0.0) *
                    (_to_float(row[idx["Capacity_Ah"]]) or 0.0)
                    if "Discharge_MaxConstant_Crate" in idx else 0.0
                ),
                taux_swelling_pct = _SWELLING_BY_CHEMISTRY.get(
                    _clean_chemistry(
                        row[idx["Chemistry_Detail"]] if "Chemistry_Detail" in idx else None,
                        row[idx["Chemistry"]] if "Chemistry" in idx else None,
                    ),
                    _SWELLING_DEFAULT,
                ),

                fabricant              = str(row[idx["Company"]]).strip() if "Company" in idx and not _is_null(row[idx["Company"]]) else None,
                chimie                 = _clean_chemistry(
                    row[idx["Chemistry_Detail"]] if "Chemistry_Detail" in idx else None,
                    row[idx["Chemistry"]] if "Chemistry" in idx else None,
                ),
                cycle_life             = _to_int(row[idx["CycleLife"]]) if "CycleLife" in idx else None,
                dod_reference_pct      = dod_raw,
                c_rate_max_discharge   = _to_float(row[idx["Discharge_MaxConstant_Crate"]]) if "Discharge_MaxConstant_Crate" in idx else None,
                c_rate_max_charge      = _to_float(row[idx["Charge_MaxConstant_Crate"]]) if "Charge_MaxConstant_Crate" in idx else None,
                eol_capacity_pct   = _to_float(row[idx["LastCycle_ResidualCapacity_%"]]) if "LastCycle_ResidualCapacity_%" in idx else None,
                cutoff_voltage_v   = _to_float(row[idx["CutOff_Voltage_Volt"]]) if "CutOff_Voltage_Volt" in idx else None,
                temp_min_c         = _to_float(row[idx["Temp_Min_C"]]) if "Temp_Min_C" in idx else None,
                temp_max_c         = _to_float(row[idx["Temp_Max_C"]]) if "Temp_Max_C" in idx else None,
                temp_max_charge_c  = _to_float(row[idx["Temp_Max_Charge_C"]]) if "Temp_Max_Charge_C" in idx else None,
                v_charge_max       = _to_float(row[idx["VCharge_Max_V"]]) if "VCharge_Max_V" in idx else None,
            ))
            count += 1

        db.commit()
    except Exception:
        db.rollback()
        raise

    return count


def _import_legacy(ws, db: Session) -> int:
    """Import from curated-format Excel (direct column names: 'nom', 'chimie', etc.)."""
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    missing = REQUIRED_COLUMNS - set(headers)
    if missing:
        raise ValueError(f"Missing columns in Excel file: {', '.join(sorted(missing))}")

    idx = {h: i for i, h in enumerate(headers)}

    def _get(row, col):
        return row[idx[col]] if col in idx else None

    db.query(Cellule).delete()

    count = 0
    try:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in row):
                continue
            cell_type = str(_get(row, "type_cellule"))
            db.add(Cellule(
                nom               = str(_get(row, "nom")),
                longueur_mm       = _to_float(_get(row, "longueur_mm")) or 0.0,
                largeur_mm        = _to_float(_get(row, "largeur_mm"))  or 0.0,
                hauteur_mm        = _to_float(_get(row, "hauteur_mm"))  or 0.0,
                masse_g           = _to_float(_get(row, "masse_g"))     or 0.0,
                tension_nominale  = _to_float(_get(row, "tension_nominale")) or 0.0,
                capacite_ah       = _to_float(_get(row, "capacite_ah")) or 0.0,
                courant_max_a     = _to_float(_get(row, "courant_max_a")) or 0.0,
                type_cellule      = cell_type,
                taux_swelling_pct = _to_float(_get(row, "taux_swelling_pct")) or 0.0,
                # Cylindrical: longueur_mm holds the diameter value
                diameter_mm       = _to_float(_get(row, "longueur_mm")) if cell_type == "Cylindrical" else None,
                # Extended fields — all present in battery_cells_curated.xlsx
                fabricant            = str(_get(row, "fabricant")).strip() if not _is_null(_get(row, "fabricant")) else None,
                chimie               = str(_get(row, "chimie")).strip()    if not _is_null(_get(row, "chimie"))    else None,
                cycle_life           = _to_int(_get(row, "cycle_life")),
                dod_reference_pct    = _to_float(_get(row, "dod_reference_pct")),
                c_rate_max_discharge = _to_float(_get(row, "c_rate_max_discharge")),
                c_rate_max_charge    = _to_float(_get(row, "c_rate_max_charge")),
                eol_capacity_pct     = _to_float(_get(row, "eol_capacity_pct")),
                cutoff_voltage_v     = _to_float(_get(row, "cutoff_voltage_v")),
                v_charge_max         = _to_float(_get(row, "v_charge_max")),
                temp_min_c           = _to_float(_get(row, "temp_min_c")),
                temp_max_c           = _to_float(_get(row, "temp_max_c")),
                temp_max_charge_c    = _to_float(_get(row, "temp_max_charge_c")),
            ))
            count += 1

        db.commit()
    except Exception:
        db.rollback()
        raise

    return count


def _import_workbook(wb, db: Session) -> int:
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    if _is_extended_format(headers):
        return _import_extended(ws, db)
    return _import_legacy(ws, db)


def import_from_bytes(file_bytes: bytes, db: Session) -> int:
    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    return _import_workbook(wb, db)


def import_from_path(path: str, db: Session) -> int:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    return _import_workbook(wb, db)


# Legacy column order — used when appending to a curated-format Excel
_LEGACY_HEADERS = [
    "nom", "longueur_mm", "largeur_mm", "hauteur_mm", "diameter_mm",
    "masse_g", "tension_nominale", "capacite_ah", "courant_max_a",
    "type_cellule", "taux_swelling_pct", "fabricant", "chimie",
    "cycle_life", "dod_reference_pct", "c_rate_max_discharge",
    "c_rate_max_charge", "eol_capacity_pct", "cutoff_voltage_v",
    "temp_min_c", "temp_max_c", "temp_max_charge_c", "v_charge_max",
]

# Extended-format column → cell dict key mapping for appending
_EXT_APPEND_MAP = {
    "Product_Number":           "nom",
    "Cell_Format":              "type_cellule",
    "Capacity_Ah":              "capacite_ah",
    "NomVoltage_Volt":          "tension_nominale",
    "Weight_gr":                "masse_g",
    "Length_mm":                "longueur_mm",
    "Width_mm":                 "largeur_mm",
    "Thickness_Height_mm":      "hauteur_mm",
    "Diameter_mm":              "diameter_mm",
    "Discharge_MaxConstant_A":  "courant_max_a",
    "Company":                  "fabricant",
    "Chemistry_Detail":         "chimie",
    "CycleLife":                "cycle_life",
    "t_cycle_DOD":              "dod_reference_pct",
    "Discharge_MaxConstant_Crate": "c_rate_max_discharge",
    "Charge_MaxConstant_Crate": "c_rate_max_charge",
    "LastCycle_ResidualCapacity_%": "eol_capacity_pct",
    "CutOff_Voltage_Volt":      "cutoff_voltage_v",
    "Temp_Min_C":               "temp_min_c",
    "Temp_Max_C":               "temp_max_c",
    "Temp_Max_Charge_C":        "temp_max_charge_c",
    "VCharge_Max_V":            "v_charge_max",
}


def append_to_source_excel(cell_dict: dict) -> bool:
    """
    Append one cell row to the configured source Excel file.
    Detects legacy vs extended format and writes the right columns.
    Returns False if no source path is configured or the file doesn't exist.
    Raises on write errors so the caller can surface the message.
    """
    path = get_source_path()
    if not path:
        logging.info("append_to_source_excel: no source path configured")
        return False
    if not Path(path).exists():
        logging.warning("append_to_source_excel: source path not found: %s", path)
        return False

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    if _is_extended_format(headers):
        # Build row matching the extended headers
        inv = {v: k for k, v in _EXT_APPEND_MAP.items()}
        row = []
        for h in headers:
            field = _EXT_APPEND_MAP.get(h)
            row.append(cell_dict.get(field) if field else None)
    else:
        # Legacy format — ensure headers exist, add any missing ones
        existing = set(headers)
        for col in _LEGACY_HEADERS:
            if col not in existing:
                ws.cell(row=1, column=len(headers) + 1, value=col)
                headers.append(col)
        row = [cell_dict.get(h) for h in headers]

    ws.append(row)
    wb.save(path)
    return True
